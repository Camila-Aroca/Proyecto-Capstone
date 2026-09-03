"""Ingesta reproducible de los XLSX contextuales de Estadísticas de Género.

Los workbooks se descargan primero a ``.cache`` y sólo se publican en RAW
después de comprobar que Excel puede abrirlos y que contienen al menos una hoja.
"""

import argparse
import datetime as dt
import hashlib
import json
import logging
from pathlib import Path
import shutil
import tempfile
from typing import Any
import urllib.request

from openpyxl import load_workbook


logger = logging.getLogger(__name__)

RAW_DIR = Path("data/raw/contexto_genero")
CACHE_DIR = Path(".cache/downloads/contexto_genero")
MANIFEST_PATH = Path("data/raw/provenance_manifest.json")

SOURCES: tuple[dict[str, str], ...] = (
    {
        "id": "egresos_intento_suicida_sexo_anio",
        "filename": "egresos_intento_suicida_sexo_anio.xlsx",
        "url": "https://www.estadisticasdegenero.cl/wp-content/uploads/2023/11/Distribucion-y-razon-egresos-intento-suicida-3.xlsx",
    },
    {
        "id": "suicidio_ratio_hm_tasas_nacional_regional",
        "filename": "suicidio_ratio_hm_tasas_nacional_regional.xlsx",
        # La página oficial enlaza "sucidio" (sin la primera i); la variante
        # ortográficamente corregida devuelve 404.
        "url": "https://www.estadisticasdegenero.cl/wp-content/uploads/2021/12/Ratio-hombre-mujer-y-tasas-de-defuncion-por-sucidio-Nacional-y-regional-4.xlsx",
    },
    {
        "id": "ansiedad_depresion_sintomas_18_mas_sexo",
        "filename": "ansiedad_depresion_sintomas_18_mas_sexo.xlsx",
        "url": "https://www.estadisticasdegenero.cl/wp-content/uploads/2022/08/personas-de-18-anos-o-mas-que-presentan-sintomas-moderados-o-severos-de-ansiedad-yo-depresion-por-sexo.xlsx",
    },
    {
        "id": "prevalencia_sintomas_depresivos_sexo",
        "filename": "prevalencia_sintomas_depresivos_sexo.xlsx",
        "url": "https://www.estadisticasdegenero.cl/wp-content/uploads/2021/11/Prevalencia-de-sintomas-depresivos-Nacional-1.xlsx",
    },
)


def sha256_file(path: Path) -> str:
    """Devuelve el SHA256 de un archivo sin cargarlo por completo en memoria."""
    digest = hashlib.sha256()
    with path.open("rb") as file_handle:
        for chunk in iter(lambda: file_handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_workbook(path: Path) -> None:
    """Verifica que un XLSX sea legible y posea al menos una hoja no vacía."""
    if not path.exists() or path.stat().st_size == 0:
        raise ValueError(f"XLSX inexistente o vacío: {path.as_posix()}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if not workbook.sheetnames:
            raise ValueError(f"XLSX sin hojas: {path.as_posix()}")
        if not any(sheet.max_row > 0 and sheet.max_column > 0 for sheet in workbook.worksheets):
            raise ValueError(f"XLSX sin celdas: {path.as_posix()}")
    finally:
        workbook.close()


def _load_manifest() -> list[dict[str, Any]]:
    if not MANIFEST_PATH.exists():
        return []
    with MANIFEST_PATH.open(encoding="utf-8") as file_handle:
        payload = json.load(file_handle)
    if not isinstance(payload, list):
        raise ValueError("El manifest de provenance debe ser una lista JSON.")
    return payload


def append_provenance(
    source: dict[str, str], snapshot_path: Path, raw_path: Path, downloaded_at: str
) -> None:
    """Agrega un snapshot al manifest; nunca reemplaza entradas históricas."""
    manifest = _load_manifest()
    manifest.append(
        {
            "source_id": source["id"],
            "source_url": source["url"],
            "filename": raw_path.name,
            "raw_path": raw_path.as_posix(),
            "downloaded_at": downloaded_at,
            "file_size": snapshot_path.stat().st_size,
            "sha256": sha256_file(snapshot_path),
        }
    )
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", delete=False, dir=MANIFEST_PATH.parent,
        prefix=f"{MANIFEST_PATH.name}.", suffix=".tmp",
    ) as temporary:
        json.dump(manifest, temporary, indent=2, ensure_ascii=False)
        temporary_path = Path(temporary.name)
    temporary_path.replace(MANIFEST_PATH)


def download_source(source: dict[str, str], force: bool = False) -> bool:
    """Descarga y publica un RAW validado. Devuelve ``True`` si hubo descarga."""
    raw_path = RAW_DIR / source["filename"]
    if not force and raw_path.exists():
        validate_workbook(raw_path)
        logger.info("[SKIP] RAW válido ya disponible: %s", raw_path.as_posix())
        return False

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    downloaded_at = dt.datetime.now(dt.timezone.utc).isoformat()
    with tempfile.NamedTemporaryFile(
        delete=False, dir=CACHE_DIR, prefix=f"{source['id']}.", suffix=".xlsx"
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        request = urllib.request.Request(
            source["url"], headers={"User-Agent": "curl/8.0.1"}
        )
        with urllib.request.urlopen(request, timeout=120) as response, temporary_path.open("wb") as target:
            shutil.copyfileobj(response, target, length=1024 * 1024)
        validate_workbook(temporary_path)
        # Registrar el hash del snapshot validado antes de publicar el RAW.
        append_provenance(source, temporary_path, raw_path, downloaded_at)
        temporary_path.replace(raw_path)
        logger.info("RAW publicado: %s", raw_path.as_posix())
        return True
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def run_downloads(force: bool = False) -> dict[str, bool]:
    """Descarga los cuatro snapshots contextuales requeridos."""
    return {source["id"]: download_source(source, force=force) for source in SOURCES}


def main() -> None:
    parser = argparse.ArgumentParser(description="Descarga XLSX contextuales de Estadísticas de Género.")
    parser.add_argument("--force", action="store_true", help="Redescarga explícitamente los snapshots RAW.")
    args = parser.parse_args()
    run_downloads(force=args.force)


if __name__ == "__main__":
    main()
