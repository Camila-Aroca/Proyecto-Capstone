"""Normaliza los cuatro cuadros contextuales de Estadísticas de Género.

Cada fuente conserva su propio Parquet porque sus indicadores, períodos y
granularidades no son semánticamente intercambiables. Esta capa sólo deshace la
maquetación de Excel (títulos, encabezados múltiples y notas); no recalcula ni
completa valores publicados.
"""

import argparse
import logging
from pathlib import Path
import tempfile
from typing import Any, Iterable

from openpyxl import load_workbook
import pyarrow as pa
import pyarrow.parquet as pq


logger = logging.getLogger(__name__)

RAW_DIR = Path("data/raw/contexto_genero")
PROCESSED_DIR = Path("data/processed/contexto_genero")

SOURCE_FILES = {
    "egresos_intento_suicida_sexo_anio": "egresos_intento_suicida_sexo_anio.xlsx",
    "suicidio_ratio_hm_tasas_nacional_regional": "suicidio_ratio_hm_tasas_nacional_regional.xlsx",
    "ansiedad_depresion_sintomas_18_mas_sexo": "ansiedad_depresion_sintomas_18_mas_sexo.xlsx",
    "prevalencia_sintomas_depresivos_sexo": "prevalencia_sintomas_depresivos_sexo.xlsx",
}

OUTPUT_FILES = {
    "egresos_intento_suicida_sexo_anio": "egresos_intento_suicida_sexo_anio.parquet",
    "suicidio_ratio_hm_tasas_nacional_regional": "suicidio_ratio_hm_tasas_nacional_regional.parquet",
    "ansiedad_depresion_sintomas_18_mas_sexo": "ansiedad_depresion_sintomas_18_mas_sexo.parquet",
    "prevalencia_sintomas_depresivos_sexo": "prevalencia_sintomas_depresivos_sexo.parquet",
}

SCHEMA = pa.schema([
    ("source_id", pa.string()),
    ("source_sheet", pa.string()),
    ("geography_level", pa.string()),
    ("geography", pa.string()),
    ("region_code", pa.int32()),
    ("period", pa.string()),
    ("year", pa.int32()),
    ("sex", pa.string()),
    ("indicator", pa.string()),
    ("value", pa.float64()),
    ("value_text", pa.string()),
    ("unit", pa.string()),
])


def _value_fields(value: Any) -> tuple[float | None, str | None]:
    """Separa números de símbolos publicados sin convertirlos a missing."""
    if not isinstance(value, bool) and isinstance(value, (int, float)):
        return float(value), None
    if isinstance(value, str) and value.strip():
        return None, value
    raise ValueError(f"Valor vacío o no soportado en una celda de datos: {value!r}")


def _record(
    source_id: str,
    sheet: str,
    geography_level: str,
    geography: str,
    region_code: int | None,
    period: str,
    year: int | None,
    sex: str,
    indicator: str,
    value: Any,
    unit: str,
) -> dict[str, Any]:
    numeric_value, text_value = _value_fields(value)
    return {
        "source_id": source_id,
        "source_sheet": sheet,
        "geography_level": geography_level,
        "geography": geography,
        "region_code": region_code,
        "period": period,
        "year": year,
        "sex": sex,
        "indicator": indicator,
        "value": numeric_value,
        "value_text": text_value,
        "unit": unit,
    }


def _workbook(source_id: str):
    path = RAW_DIR / SOURCE_FILES[source_id]
    if not path.exists():
        raise FileNotFoundError(f"RAW requerido no encontrado: {path.as_posix()}")
    return load_workbook(path, read_only=True, data_only=True)


def parse_egresos_intento_suicida() -> list[dict[str, Any]]:
    """Extrae la tabla nacional del cuadro MINSAL de egresos por intento."""
    source_id = "egresos_intento_suicida_sexo_anio"
    workbook = _workbook(source_id)
    try:
        sheet = workbook["NACIONAL"]
        records: list[dict[str, Any]] = []
        definitions = (
            (1, "Total", "egresos_hospitalarios_intento_suicida", "N"),
            (2, "Hombres", "egresos_hospitalarios_intento_suicida", "N"),
            (3, "Mujeres", "egresos_hospitalarios_intento_suicida", "N"),
            (4, "Hombres", "distribucion_egresos_intento_suicida", "%"),
            (5, "Mujeres", "distribucion_egresos_intento_suicida", "%"),
            (6, "Mujeres/Hombres", "razon_egresos_intento_suicida", "razon"),
        )
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not isinstance(row[0], int):
                continue
            for index, sex, indicator, unit in definitions:
                records.append(_record(
                    source_id, "NACIONAL", "nacional", "Nacional", None,
                    str(row[0]), row[0], sex, indicator, row[index], unit,
                ))
        return records
    finally:
        workbook.close()


def parse_suicidio_ratio_tasas() -> list[dict[str, Any]]:
    """Extrae por separado observaciones nacionales y regionales del INE."""
    source_id = "suicidio_ratio_hm_tasas_nacional_regional"
    workbook = _workbook(source_id)
    try:
        records: list[dict[str, Any]] = []
        definitions = (
            (1, "Total", "defunciones_suicidio", "N"),
            (2, "Hombres", "defunciones_suicidio", "N"),
            (3, "Mujeres", "defunciones_suicidio", "N"),
            (4, "Total", "poblacion", "N"),
            (5, "Hombres", "poblacion", "N"),
            (6, "Mujeres", "poblacion", "N"),
            (7, "Total", "tasa_mortalidad_suicidio", "por_100000"),
            (8, "Hombres", "tasa_mortalidad_suicidio", "por_100000"),
            (9, "Mujeres", "tasa_mortalidad_suicidio", "por_100000"),
            (10, "Hombres/Mujeres", "ratio_defunciones_suicidio", "razon"),
        )
        for row in workbook["NACIONAL"].iter_rows(min_row=4, values_only=True):
            if not isinstance(row[0], int):
                continue
            for index, sex, indicator, unit in definitions:
                records.append(_record(
                    source_id, "NACIONAL", "nacional", "Nacional", None,
                    str(row[0]), row[0], sex, indicator, row[index], unit,
                ))
        regional_definitions = tuple(
            (index + 2, sex, indicator, unit)
            for index, sex, indicator, unit in definitions
        )
        for row in workbook["REGIONAL"].iter_rows(min_row=4, values_only=True):
            if not isinstance(row[0], int) or not isinstance(row[2], int):
                continue
            for index, sex, indicator, unit in regional_definitions:
                records.append(_record(
                    source_id, "REGIONAL", "regional", str(row[1]), row[2],
                    str(row[0]), row[0], sex, indicator, row[index], unit,
                ))
        return records
    finally:
        workbook.close()


def parse_ansiedad_depresion() -> list[dict[str, Any]]:
    """Extrae las cuatro rondas nacionales del indicador PHQ-4."""
    source_id = "ansiedad_depresion_sintomas_18_mas_sexo"
    workbook = _workbook(source_id)
    try:
        sheet = workbook["NACIONAL"]
        definitions = (
            (2, "Total", "personas_18_mas", "N"),
            (4, "Hombres", "personas_18_mas", "N"),
            (6, "Mujeres", "personas_18_mas", "N"),
            (8, "Total", "personas_18_mas_sintomas_moderados_severos_ansiedad_depresion", "N"),
            (10, "Hombres", "personas_18_mas_sintomas_moderados_severos_ansiedad_depresion", "N"),
            (12, "Mujeres", "personas_18_mas_sintomas_moderados_severos_ansiedad_depresion", "N"),
            (14, "Total", "porcentaje_sintomas_moderados_severos_ansiedad_depresion", "%"),
            (16, "Hombres", "porcentaje_sintomas_moderados_severos_ansiedad_depresion", "%"),
            (18, "Mujeres", "porcentaje_sintomas_moderados_severos_ansiedad_depresion", "%"),
            (19, "Mujeres-Hombres", "brecha_genero_sintomas_moderados_severos_ansiedad_depresion", "pp"),
        )
        records: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=5, max_col=20, values_only=True):
            if not isinstance(row[0], str) or not isinstance(row[2], (int, float)):
                continue
            for index, sex, indicator, unit in definitions:
                records.append(_record(
                    source_id, "NACIONAL", "nacional", "Nacional", None,
                    row[0], None, sex, indicator, row[index], unit,
                ))
        return records
    finally:
        workbook.close()


def parse_prevalencia_depresivos() -> list[dict[str, Any]]:
    """Extrae las prevalencias nacionales ENS, preservando el período publicado."""
    source_id = "prevalencia_sintomas_depresivos_sexo"
    workbook = _workbook(source_id)
    try:
        sheet = workbook["NACIONAL"]
        records: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=6, max_col=5, values_only=True):
            if not isinstance(row[0], (int, str)) or not isinstance(row[2], (int, float)):
                continue
            period = str(row[0])
            year = row[0] if isinstance(row[0], int) else None
            for index, sex in ((2, "Hombres"), (4, "Mujeres")):
                records.append(_record(
                    source_id, "NACIONAL", "nacional", "Nacional", None,
                    period, year, sex, "prevalencia_sintomas_depresivos_ultimo_ano",
                    row[index], "%",
                ))
        return records
    finally:
        workbook.close()


PARSERS = {
    "egresos_intento_suicida_sexo_anio": parse_egresos_intento_suicida,
    "suicidio_ratio_hm_tasas_nacional_regional": parse_suicidio_ratio_tasas,
    "ansiedad_depresion_sintomas_18_mas_sexo": parse_ansiedad_depresion,
    "prevalencia_sintomas_depresivos_sexo": parse_prevalencia_depresivos,
}


def validate_records(records: Iterable[dict[str, Any]], source_id: str) -> list[dict[str, Any]]:
    """Valida clave natural y ausencia de valores faltantes en valores publicados."""
    materialized = list(records)
    if not materialized:
        raise ValueError(f"La normalización no produjo filas para {source_id}.")
    key_columns = ("source_id", "source_sheet", "geography_level", "geography", "region_code",
                   "period", "sex", "indicator", "unit")
    seen: set[tuple[Any, ...]] = set()
    for record in materialized:
        if record["value"] is None and record["value_text"] is None:
            raise ValueError(f"Valor faltante no permitido en {source_id}: {record}")
        key = tuple(record[column] for column in key_columns)
        if key in seen:
            raise ValueError(f"Duplicado según clave natural en {source_id}: {key}")
        seen.add(key)
    return materialized


def write_records(source_id: str, records: Iterable[dict[str, Any]]) -> Path:
    """Escribe Parquet de manera atómica, sin alterar un output válido si falla."""
    validated = validate_records(records, source_id)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    output_path = PROCESSED_DIR / OUTPUT_FILES[source_id]
    with tempfile.NamedTemporaryFile(
        delete=False, dir=PROCESSED_DIR, prefix=f"{output_path.name}.", suffix=".tmp"
    ) as temporary:
        temporary_path = Path(temporary.name)
    try:
        table = pa.Table.from_pylist(validated, schema=SCHEMA)
        pq.write_table(table, temporary_path, compression="snappy")
        temporary_path.replace(output_path)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise
    return output_path


def run_normalization() -> dict[str, Path]:
    """Genera un Parquet independiente por fuente contextual."""
    outputs: dict[str, Path] = {}
    for source_id, parser in PARSERS.items():
        outputs[source_id] = write_records(source_id, parser())
        logger.info("Normalizada %s: %s", source_id, outputs[source_id].as_posix())
    return outputs


def main() -> None:
    parser = argparse.ArgumentParser(description="Normaliza XLSX contextuales de Estadísticas de Género.")
    parser.add_argument("--force", action="store_true", help="Aceptado para interoperar con el orquestador.")
    parser.parse_args()
    run_normalization()


if __name__ == "__main__":
    main()
