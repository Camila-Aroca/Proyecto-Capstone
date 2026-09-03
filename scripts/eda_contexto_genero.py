"""EDA reproducible de los cuatro cuadros contextuales de Estadísticas de Género."""

import argparse
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pyarrow.parquet as pq

from src.data.normalize_gender_statistics import (
    OUTPUT_FILES,
    PROCESSED_DIR,
    RAW_DIR,
    SOURCE_FILES,
)


REPORT_PATH = Path("reports/eda/eda_contexto_genero_estadisticas_genero.md")

SOURCE_LAYOUTS: dict[str, dict[str, str]] = {
    "egresos_intento_suicida_sexo_anio": {
        "headers": "NACIONAL: fila 3; datos desde fila 4.",
        "grain": "año nacional × medida/sexo.",
        "notes": "Título en fila 1; fuente y notas al final de la hoja NACIONAL. PRESENTACIÓN contiene texto metodológico.",
    },
    "suicidio_ratio_hm_tasas_nacional_regional": {
        "headers": "NACIONAL y REGIONAL: fila 3; datos desde fila 4.",
        "grain": "año × nivel geográfico (nacional o región) × medida/sexo.",
        "notes": "Título en fila 1; notas de cobertura y tasas al final de cada cuadro. PRESENTACIÓN contiene texto metodológico.",
    },
    "ansiedad_depresion_sintomas_18_mas_sexo": {
        "headers": "NACIONAL: encabezado multinivel en filas 3–4; datos desde fila 5.",
        "grain": "período nacional × medida/sexo.",
        "notes": "Título en fila 1; fuente y notas 1–3 al final. PRESENTACIÓN contiene texto metodológico.",
    },
    "prevalencia_sintomas_depresivos_sexo": {
        "headers": "NACIONAL: encabezado multinivel en filas 4–5; datos desde fila 6.",
        "grain": "período nacional × sexo.",
        "notes": "Título en fila 1; fuente y notas 1–2 al final. PRESENTACIÓN contiene texto metodológico.",
    },
}


def workbook_structure(source_id: str) -> list[dict[str, Any]]:
    """Obtiene nombres, dimensiones declaradas y título de las hojas del RAW."""
    path = RAW_DIR / SOURCE_FILES[source_id]
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return [
            {
                "sheet": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "title": sheet.cell(1, 1).value,
            }
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def profile_processed(source_id: str) -> dict[str, Any]:
    """Calcula perfil directamente desde el Parquet normalizado de una fuente."""
    rows = pq.read_table(PROCESSED_DIR / OUTPUT_FILES[source_id]).to_pylist()
    natural_key = (
        "source_id", "source_sheet", "geography_level", "geography", "region_code",
        "period", "sex", "indicator", "unit",
    )
    keys = [tuple(row[column] for column in natural_key) for row in rows]
    return {
        "rows": len(rows),
        "columns": list(rows[0]) if rows else [],
        "years": sorted({row["year"] for row in rows if row["year"] is not None}),
        "periods_without_year": sorted({row["period"] for row in rows if row["year"] is None}),
        "geographies": sorted({row["geography"] for row in rows}),
        "geography_levels": sorted({row["geography_level"] for row in rows}),
        "sexes": sorted({row["sex"] for row in rows}),
        "indicators": sorted({row["indicator"] for row in rows}),
        "units": sorted({row["unit"] for row in rows}),
        "missing_numeric": sum(row["value"] is None for row in rows),
        "text_values": Counter(row["value_text"] for row in rows if row["value_text"] is not None),
        "duplicates": len(keys) - len(set(keys)),
        "numeric_min": min(row["value"] for row in rows if row["value"] is not None),
        "numeric_max": max(row["value"] for row in rows if row["value"] is not None),
    }


def _inline(values: list[Any]) -> str:
    return ", ".join(str(value) for value in values) if values else "No aplica"


def build_report() -> str:
    """Construye el reporte, distinguiendo observación de interpretación."""
    sections = [
        "# EDA — fuentes contextuales de Estadísticas de Género",
        "",
        "## Alcance",
        "",
        "Este reporte perfila cuatro XLSX oficiales como fuentes contextuales para el análisis de hospitalización por sexo. No realiza cruces, uniones ni inferencias respecto de Egresos DEIS o Urgencias.",
        "",
        "## Hechos observados",
    ]
    for source_id in SOURCE_FILES:
        structure = workbook_structure(source_id)
        profile = profile_processed(source_id)
        raw_structure = _inline([
            "{} ({} filas × {} columnas)".format(
                item["sheet"], item["rows"], item["columns"]
            )
            for item in structure
        ])
        sections.extend([
            "",
            f"### `{source_id}`",
            "",
            f"- Estructura RAW: {raw_structure}.",
            f"- Encabezados y grano: {SOURCE_LAYOUTS[source_id]['headers']} Grano normalizado: {SOURCE_LAYOUTS[source_id]['grain']}",
            f"- Títulos/notas: {SOURCE_LAYOUTS[source_id]['notes']}",
            f"- Output: `{PROCESSED_DIR.as_posix()}/{OUTPUT_FILES[source_id]}`; {profile['rows']} filas; columnas: `{_inline(profile['columns'])}`.",
            f"- Cobertura temporal: años {_inline(profile['years'])}; períodos sin año único: {_inline(profile['periods_without_year'])}.",
            f"- Cobertura geográfica: niveles {_inline(profile['geography_levels'])}; geografías {_inline(profile['geographies'])}.",
            f"- Categorías de sexo: {_inline(profile['sexes'])}. Indicadores: {_inline(profile['indicators'])}. Unidades: {_inline(profile['units'])}.",
            f"- Calidad: missing numérico {profile['missing_numeric']}; valores textuales publicados `{dict(profile['text_values'])}`; duplicados por clave natural {profile['duplicates']}; rango numérico [{profile['numeric_min']}, {profile['numeric_max']}].",
        ])

    sections.extend([
        "",
        "## Anomalías y limitaciones observadas",
        "",
        "- El cuadro PHQ-4 declara 16.380 columnas en `NACIONAL`, aunque la tabla con contenido usa las columnas A:T; las columnas adicionales son una anomalía de estructura/formatación del XLSX. La normalización usa explícitamente A:T y no modifica el RAW.",
        "- El título del cuadro de egresos por intentos suicidas declara años 2006–2024, pero sus filas de datos incluyen 2025. El EDA y el Parquet conservan la fila observada y documentan la discrepancia.",
        "- En el cuadro regional de suicidio existe al menos un ratio publicado como símbolo `-`; se conserva en `value_text` y no se transforma en cero ni se imputa.",
        "- Los cuadros nacionales no proporcionan desagregación regional; los períodos `2009-10`, `2016-17` y las rondas de Encuesta Social Covid-19 no se convierten a un año artificial.",
        "- Las categorías `Total`, razones y brechas son medidas publicadas, no categorías de personas. Los indicadores se mantienen separados por fuente y no deben interpretarse como observaciones individuales ni vincularse a Egresos DEIS.",
        "",
        "## Interpretación acotada",
        "",
        "Los cuatro cuadros permiten contextualizar diferencias publicadas por sexo, con cobertura temporal y geográfica heterogénea. No es posible determinar asociaciones con duración de estadía ni con egresos individuales usando estos datos disponibles.",
    ])
    return "\n".join(sections) + "\n"


def run_eda() -> Path:
    """Genera el EDA Markdown en su ubicación canónica."""
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(build_report(), encoding="utf-8")
    return REPORT_PATH


def main() -> None:
    parser = argparse.ArgumentParser(description="EDA de fuentes contextuales de Estadísticas de Género.")
    parser.add_argument("--force", action="store_true", help="Aceptado para interoperar con el orquestador.")
    parser.parse_args()
    print(run_eda().as_posix())


if __name__ == "__main__":
    main()
